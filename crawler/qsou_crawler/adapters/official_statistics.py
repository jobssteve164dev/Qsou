"""Shared parser for official statistical releases and downloadable tables."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from html.parser import HTMLParser
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Any, Optional, Sequence
from urllib.parse import urlsplit

from .base import (
    DocumentReference,
    ResponsePayload,
    SourceAdapter,
    normalize_text,
    parse_html,
    parse_published_at,
)


MAX_STRUCTURED_CELLS = 100_000


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return normalize_text(value)


class _HTMLTableParser(HTMLParser):
    def __init__(self, container_patterns: Sequence[str]) -> None:
        super().__init__(convert_charrefs=True)
        self.container_patterns = tuple(container_patterns)
        self.tables: list[list[list[str]]] = []
        self._container_depth = 0
        self._stack: list[tuple[str, bool]] = []
        self._table_depth = 0
        self._row: Optional[list[str]] = None
        self._cell: Optional[list[str]] = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        attributes = {key.lower(): value or "" for key, value in attrs if key}
        selector = normalize_text(
            " ".join((attributes.get("id", ""), attributes.get("class", "")))
        )
        is_container = bool(selector) and any(
            re.search(pattern, selector, re.IGNORECASE)
            for pattern in self.container_patterns
        )
        self._stack.append((tag, is_container))
        if is_container:
            self._container_depth += 1
        if tag == "table" and self._container_depth:
            self._table_depth += 1
            if self._table_depth == 1:
                self.tables.append([])
        elif tag == "tr" and self._table_depth:
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            value = normalize_text(data)
            if value:
                self._cell.append(value)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"th", "td"} and self._cell is not None and self._row is not None:
            self._row.append(normalize_text(" ".join(self._cell)))
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table_depth:
            if any(self._row):
                self.tables[-1].append(self._row)
            self._row = None
        elif tag == "table" and self._table_depth:
            self._table_depth -= 1
        for index in range(len(self._stack) - 1, -1, -1):
            if self._stack[index][0] != tag:
                continue
            removed = self._stack[index:]
            del self._stack[index:]
            self._container_depth = max(
                0,
                self._container_depth - sum(1 for _, matched in removed if matched),
            )
            break


class OfficialStatisticsAdapter(SourceAdapter):
    """Normalize one official release without treating derived rows as raw evidence."""

    document_type = "statistical_release"
    content_container_patterns = (
        r"trs_editor",
        r"detail_content",
        r"article[-_ ]?content",
        r"content[-_ ]?detail",
        r"zoom",
    )

    def parse_document(
        self,
        response: ResponsePayload,
        reference: DocumentReference,
    ) -> Optional[dict[str, Any]]:
        suffix = PurePosixPath(urlsplit(response.url).path).suffix.lower()
        content_type = response.content_type.lower()
        structured_data: dict[str, Any] = {
            "schema": "qsou.statistical_release.v1",
            "tables": [],
        }
        published_at = reference.published_at

        if suffix in {".xlsx", ".xls"} or "spreadsheet" in content_type or "excel" in content_type:
            tables = self._spreadsheet_tables(response.body, suffix)
            structured_data["tables"] = tables
            content = self._render_tables(tables)
            title = normalize_text(reference.title) or self.reference_id(response.url)
            extraction = "official_spreadsheet"
        elif suffix == ".csv" or "text/csv" in content_type:
            rows = list(csv.reader(StringIO(response.text)))
            tables = self._bounded_tables([{"name": "data", "rows": rows}])
            structured_data["tables"] = tables
            content = self._render_tables(tables)
            title = normalize_text(reference.title) or self.reference_id(response.url)
            extraction = "official_csv"
        elif suffix == ".pdf" or "pdf" in content_type:
            content = normalize_text(self._extract_pdf(response.body))
            title = normalize_text(reference.title) or self.reference_id(response.url)
            extraction = "official_pdf"
        else:
            parser = parse_html(response.text, self.content_container_patterns)
            title = normalize_text(
                parser.meta.get("og:title")
                or " ".join(parser.heading_parts)
                or reference.title
                or " ".join(parser.title_parts)
            )
            blocks = parser.content_blocks or parser.semantic_blocks
            content = normalize_text("\n".join(dict.fromkeys(blocks)))
            table_parser = _HTMLTableParser(self.content_container_patterns)
            table_parser.feed(response.text)
            tables = self._bounded_tables(
                [
                    {"name": f"table_{index + 1}", "rows": rows}
                    for index, rows in enumerate(table_parser.tables)
                    if rows
                ]
            )
            structured_data["tables"] = tables
            rendered = self._render_tables(tables)
            if rendered and rendered not in content:
                content = normalize_text(f"{content}\n{rendered}")
            published_at = published_at or parse_published_at(parser, response.text)
            extraction = "official_html_release"

        content = normalize_text(f"{title}\n{content}")
        attribution = normalize_text(self.source.get("attribution"))
        if attribution:
            content = normalize_text(f"{content}\n{attribution}")
        if len(title) < 4 or len(content) < 50:
            return None
        structured_data["table_count"] = len(structured_data["tables"])
        return {
            "source_document_id": reference.source_document_id or self.reference_id(response.url),
            "type": self.document_type,
            "title": title[:500],
            "content": content[:500_000],
            "url": response.url,
            "source": self.source.get("source_name", self.source_id),
            "source_id": self.source_id,
            "source_published_at": published_at,
            "parser_version": f"{self.adapter_id}/{self.version}",
            "structured_data": structured_data,
            "metadata": {
                "adapter_id": self.adapter_id,
                "adapter_version": self.version,
                "extraction": extraction,
                "content_granularity": "official_statistical_release",
                "rights_status": self.source.get("rights_status"),
                **dict(reference.metadata),
            },
        }

    @staticmethod
    def _spreadsheet_tables(body: bytes, suffix: str) -> list[dict[str, Any]]:
        tables: list[dict[str, Any]] = []
        if suffix == ".xls":
            try:
                import xlrd
            except ImportError as exc:
                raise ValueError("解析官方 XLS 需要 xlrd") from exc
            workbook = xlrd.open_workbook(file_contents=body, on_demand=True)
            for sheet in workbook.sheets():
                tables.append(
                    {
                        "name": normalize_text(sheet.name),
                        "rows": [
                            [_cell_value(sheet.cell_value(row, column)) for column in range(sheet.ncols)]
                            for row in range(sheet.nrows)
                        ],
                    }
                )
            workbook.release_resources()
        else:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("解析官方 XLSX 需要 openpyxl") from exc
            workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                tables.append(
                    {
                        "name": normalize_text(sheet.title),
                        "rows": [[_cell_value(value) for value in row] for row in sheet.iter_rows(values_only=True)],
                    }
                )
            workbook.close()
        return OfficialStatisticsAdapter._bounded_tables(tables)

    @staticmethod
    def _bounded_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        cells = 0
        for table in tables:
            rows: list[list[str]] = []
            for raw_row in table.get("rows", []):
                row = [_cell_value(value) for value in raw_row]
                while row and not row[-1]:
                    row.pop()
                if not any(row):
                    continue
                cells += len(row)
                if cells > MAX_STRUCTURED_CELLS:
                    raise ValueError(
                        f"官方统计表超过单文档结构化上限: {MAX_STRUCTURED_CELLS} cells"
                    )
                rows.append(row)
            if rows:
                normalized.append({"name": normalize_text(table.get("name")), "rows": rows})
        return normalized

    @staticmethod
    def _render_tables(tables: list[dict[str, Any]]) -> str:
        lines: list[str] = []
        for table in tables:
            name = normalize_text(table.get("name"))
            if name:
                lines.append(name)
            lines.extend("\t".join(row) for row in table.get("rows", []))
        return "\n".join(lines)
